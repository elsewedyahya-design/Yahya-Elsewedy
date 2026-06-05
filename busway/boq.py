"""Price-breakdown (BOQ) ingestion (blueprint §4.4, Task 7).

Parses a pasted-from-Excel table (tab/multi-space separated) OR an .xlsx file.
Matching is by HEADER NAME, not column position, so the common case is
frictionless and a reordered sheet still works. Anything ambiguous is FLAGGED
for confirmation rather than silently shipped wrong:
  * missing required column            -> error
  * unrecognised / extra column        -> warning
  * line-item total != qty x unit_price -> warning
  * leftover "xxx" placeholder          -> warning
Group sub-header rows (a label with no qty/price) are kept but marked, so the
generated table can group line items by product.
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field

# canonical field -> accepted header synonyms (normalised, lowercase)
SYNONYMS: dict[str, set[str]] = {
    "item": {"item", "item no", "no", "sr", "sr no", "s no", "#"},
    "type": {"type", "product", "product type", "category"},
    "description": {"description", "desc", "details", "scope"},
    "qty": {"qty", "quantity", "qnty", "qty.", "q'ty"},
    "unit": {"unit", "uom", "u/m", "units"},
    "unit_price": {"unit price", "unit rate", "rate", "price", "unit cost",
                   "unitprice"},
    "total_price": {"total price", "total", "amount", "line total",
                    "total cost", "totalprice"},
}
REQUIRED = {"description", "qty", "unit_price", "total_price"}


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip().lower()).strip(" .:")


def _to_float(v: str) -> float | None:
    s = str(v).strip().replace(",", "").replace("$", "")
    if s == "" or s.lower() in {"xxx", "tbd", "n/a", "-"}:
        return None
    try:
        return float(s)
    except ValueError:
        return None


@dataclass
class BOQResult:
    rows: list[dict] = field(default_factory=list)     # line items + group rows
    columns: dict[str, int] = field(default_factory=dict)  # field -> col index
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)

    @property
    def ok(self) -> bool:
        return not self.errors


def _split_line(line: str) -> list[str]:
    # Prefer tabs (Excel copy); fall back to runs of 2+ spaces.
    if "\t" in line:
        return [c.strip() for c in line.split("\t")]
    return [c.strip() for c in re.split(r" {2,}", line.strip())]


def _map_headers(cells: list[str]) -> tuple[dict[str, int], list[str]]:
    """Map header cells -> canonical field indices. Returns (mapping, warnings)."""
    mapping: dict[str, int] = {}
    warnings: list[str] = []
    for idx, cell in enumerate(cells):
        key = _norm(cell)
        if not key:
            continue
        matched = next((f for f, syns in SYNONYMS.items() if key in syns), None)
        if matched and matched not in mapping:
            mapping[matched] = idx
        elif matched:
            warnings.append(f"duplicate column for '{matched}' at position {idx + 1}")
        else:
            warnings.append(f"unrecognised column '{cell}' (position {idx + 1}) — ignored")
    return mapping, warnings


def parse_rows(table: list[list[str]]) -> BOQResult:
    """Parse a list of cell-rows (first row = header)."""
    res = BOQResult()
    if not table:
        res.errors.append("empty BOQ")
        return res
    header, *body = table
    mapping, warns = _map_headers(header)
    res.columns = mapping
    res.warnings.extend(warns)

    missing = REQUIRED - set(mapping)
    if missing:
        res.errors.append(
            "missing required column(s): " + ", ".join(sorted(missing))
            + " — check the header names")
        return res

    def cell(row: list[str], field_: str) -> str:
        i = mapping.get(field_)
        return row[i].strip() if i is not None and i < len(row) else ""

    for n, row in enumerate(body, start=2):
        if not any(c.strip() for c in row):
            continue
        desc = cell(row, "description") or cell(row, "type")
        qty = _to_float(cell(row, "qty"))
        up = _to_float(cell(row, "unit_price"))
        tp = _to_float(cell(row, "total_price"))

        rec = {
            "item": cell(row, "item"),
            "type": cell(row, "type"),
            "description": desc,
            "qty": cell(row, "qty"),
            "unit": cell(row, "unit"),
            "unit_price": cell(row, "unit_price"),
            "total_price": cell(row, "total_price"),
        }

        # group sub-header: a label with no numeric qty AND no price
        if qty is None and up is None and tp is None and desc:
            rec["is_group_header"] = True
            res.rows.append(rec)
            continue

        rec["is_group_header"] = False
        # leftover placeholder check
        if any("xxx" in str(v).lower() for v in rec.values()):
            res.warnings.append(f"row {n}: leftover 'xxx' placeholder — {desc!r}")
        # total sanity check
        if qty is not None and up is not None and tp is not None:
            if abs(qty * up - tp) > max(0.01, 0.01 * abs(tp)):
                res.warnings.append(
                    f"row {n}: total {tp:g} != qty×unit_price "
                    f"({qty:g}×{up:g}={qty * up:g}) — {desc!r}")
        res.rows.append(rec)

    if not any(not r.get("is_group_header") for r in res.rows):
        res.warnings.append("no line items detected (only group headers?)")
    return res


def parse_pasted(text: str) -> BOQResult:
    """Parse text pasted from Excel (tab- or multi-space-separated)."""
    lines = [ln for ln in text.splitlines() if ln.strip()]
    return parse_rows([_split_line(ln) for ln in lines])


def parse_xlsx(path: str) -> BOQResult:
    """Parse the first worksheet of an .xlsx file."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    table = []
    for row in ws.iter_rows(values_only=True):
        table.append(["" if c is None else str(c) for c in row])
    wb.close()
    return parse_rows(table)
